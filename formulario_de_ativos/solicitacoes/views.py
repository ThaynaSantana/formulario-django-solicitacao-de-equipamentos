from django.shortcuts import render, redirect
from .forms import SolicitacaoForm
import openpyxl
from openpyxl.utils import get_column_letter

# Create your views here.

def save_to_excel(data):
    file_path = 'https://1drv.ms/x/c/e9d0690692d5011e/EZ8sT9-3lUhOlqfbGYasImUB7W7MEYS19uvHi9beJduj6Q?e=L1fdJQ'

    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    planilha = workbook.active

    if planilha.max_row == 1 and planilha.cell(row=1, column=1).value is None:
        # Cabeçalho da tabela
        for col_num, key in enumerate(data.keys(), 1):
            column_letter = get_column_letter(col_num)
            planilha[f'{column_letter}1']
    
    # Dados da tabela inserção
    row_num = planilha.max_row + 1
    for col_num, (key, value) in enumerate(data.items(), 1):
        column_letter = get_column_letter(col_num)
        planilha[f'{column_letter}{row_num}'] = value
    
    # Salvando planilha
    workbook.save(file_path)

def index(request):
    if request.method == 'POST':
        form = SolicitacaoForm(request.POST)
        if form.is_valid():
            form_data = form.cleaned_data
            save_to_excel(form_data)
            return redirect('sucesso.html')
    else:
            form = SolicitacaoForm()
    return render(request, 'formulario.html', {'form': form})