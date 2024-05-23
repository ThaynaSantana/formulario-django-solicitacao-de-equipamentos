# handle excel
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook


def listar_ativos_disponiveis():
    planilha = "planilhas/GESTAO_ATIVOS.xlsx"
    df = pd.read_excel(planilha, sheet_name='Inventario')
    # Filtra os ativos com status "disponível"
    ativos_disponiveis = df[df['STATUS'] == 'DISPONIVEL']['COD PATRIMONIAL'].tolist()
    return ativos_disponiveis


# Testando
ativos_disponiveis = listar_ativos_disponiveis()
ativos_tuplas = [(ativo, ativo) for ativo in ativos_disponiveis]
CODIGO_ATIVO_CHOICES = tuple(ativos_tuplas)
print(CODIGO_ATIVO_CHOICES)


def inserir_dados_solicitacao(dados_formulario):
    # Carrega a planilha existente
    planilha = "planilhas/SOLICITAÇÕES.xlsx"
    workbook = load_workbook(planilha)
    sheet = workbook.active
    next_row = sheet.max_row + 1  # Encontra a próxima linha vazia na planilha
    # data e hora da solicitação
    data_hora_envio = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    dados_formulario.append(data_hora_envio)
    # Insere os dados do formulário na planilha
    for col, value in enumerate(dados_formulario, start=1):
        sheet.cell(row=next_row, column=col, value=value)

    workbook.save(planilha)
